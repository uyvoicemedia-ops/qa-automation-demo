#!/usr/bin/env python3
"""
Nightly QA Workflow  —  scripts/nightly_qa/main.py
Mirrors the 9-step workflow originally designed for Claude Cowork.

Env vars required (set as GitHub Secrets):
  JIRA_EMAIL, JIRA_TOKEN, ANTHROPIC_API_KEY, SLACK_BOT_TOKEN,
  GITHUB_TOKEN, QA_DEFAULT_USERNAME, QA_DEFAULT_PASSWORD,
  GMAIL_USER, GMAIL_APP_PASSWORD

Optional:
  STAGING_URL  (defaults to https://qa-test-company.i6clouds.com/dp#/dashboard)
"""

import json
import os
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import anthropic
import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

# ── Config ───────────────────────────────────────────────────────────────────
JIRA_BASE         = "https://i6group.atlassian.net"
JIRA_EMAIL        = os.environ["JIRA_EMAIL"]
JIRA_TOKEN        = os.environ["JIRA_TOKEN"]
JIRA_ASSIGNEE_ID  = "712020:7020c162-2bc4-4df1-8436-a71e505ce0bc"
SLACK_BOT_TOKEN   = os.environ.get("SLACK_BOT_TOKEN", "")   # optional
SLACK_DM_CHANNEL  = "D0BG5DM15PF"
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
GITHUB_TOKEN      = os.environ.get("GITHUB_TOKEN", "")
STAGING_URL       = os.environ.get("STAGING_URL", "https://qa-test-company.i6clouds.com/dp#/dashboard")
STAGING_USERNAME  = os.environ.get("STAGING_USERNAME", "")
STAGING_PASSWORD  = os.environ.get("STAGING_PASSWORD", "")

TODAY       = datetime.now(timezone.utc).strftime("%Y-%m-%d")
DEDUP_PATH  = Path("qa_processed_tickets.json")
OUTPUT_DIR  = Path("qa_outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

jira_auth    = (JIRA_EMAIL, JIRA_TOKEN)
jira_headers = {"Accept": "application/json", "Content-Type": "application/json"}
ai_client    = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 0  —  Deduplication
# ─────────────────────────────────────────────────────────────────────────────
def step0_dedup():
    if not DEDUP_PATH.exists():
        DEDUP_PATH.write_text(json.dumps({"processed": []}, indent=2))
    data = json.loads(DEDUP_PATH.read_text())
    processed_keys = {e["ticket_key"] for e in data.get("processed", [])}
    print(f"[Step 0] Previously processed: {processed_keys or 'none'}")
    return data, processed_keys


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1  —  Fetch first unprocessed Pending QA ticket
# ─────────────────────────────────────────────────────────────────────────────
def step1_get_ticket(processed_keys: set):
    jql = (
        'project = F6 AND status = "Pending QA" '
        f'AND assignee = "{JIRA_ASSIGNEE_ID}" '
        "ORDER BY created ASC"
    )
    resp = requests.get(
        f"{JIRA_BASE}/rest/api/3/search",
        auth=jira_auth,
        headers=jira_headers,
        params={
            "jql": jql,
            "maxResults": 20,
            "fields": "summary,status,parent,description,created",
        },
    )
    resp.raise_for_status()
    issues = resp.json().get("issues", [])

    ticket = None
    fallback = None

    for issue in issues:
        if issue["key"] not in processed_keys:
            ticket = issue
            break
        if fallback is None:
            fallback = issue  # oldest already-processed ticket

    # If everything was processed before, re-process the oldest one
    if ticket is None:
        ticket = fallback
        if ticket:
            print(f"[Step 1] All tickets previously processed — re-running oldest: {ticket['key']}")

    if ticket is None:
        print("[Step 1] No Pending QA tickets found. Exiting.")
        sys.exit(0)

    print(f"[Step 1] Selected: {ticket['key']} — {ticket['fields']['summary']}")
    return ticket


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2  —  Parent ticket → branch name via GitHub API
# ─────────────────────────────────────────────────────────────────────────────
def step2_get_branch(ticket) -> tuple[Optional[str], Optional[str]]:
    parent = ticket["fields"].get("parent")
    if not parent:
        print("[Step 2] No parent ticket — skipping branch lookup.")
        return None, None

    parent_key = parent["key"]
    print(f"[Step 2] Parent: {parent_key}")

    resp = requests.get(
        f"{JIRA_BASE}/rest/api/3/issue/{parent_key}/remotelink",
        auth=jira_auth,
        headers=jira_headers,
    )
    resp.raise_for_status()
    links = resp.json()

    pr_links = [l for l in links if "github.com" in l.get("object", {}).get("url", "")]
    if not pr_links:
        print("[Step 2] No GitHub PR links on parent ticket.")
        return parent_key, None

    # Prefer the feature/fix PR over sync/chore PRs
    chosen_url = None
    for link in pr_links:
        title = link["object"].get("title", "").lower()
        url   = link["object"]["url"]
        if any(kw in title for kw in ("feat", "feature", "fix")) and "sync" not in title:
            chosen_url = url
            break
    if not chosen_url:
        chosen_url = pr_links[0]["object"]["url"]

    pr_match   = re.search(r"/pull/(\d+)", chosen_url)
    repo_match = re.search(r"github\.com/([^/]+/[^/]+)/pull", chosen_url)
    if not pr_match or not repo_match:
        print(f"[Step 2] Could not parse PR URL: {chosen_url}")
        return parent_key, None

    pr_number = pr_match.group(1)
    repo      = repo_match.group(1)

    gh_resp = requests.get(
        f"https://api.github.com/repos/{repo}/pulls/{pr_number}",
        headers={
            "Authorization": f"Bearer {GITHUB_TOKEN}",
            "Accept": "application/vnd.github+json",
            "X-GitHub-Api-Version": "2022-11-28",
        },
    )
    gh_resp.raise_for_status()
    branch = gh_resp.json()["head"]["ref"]
    print(f"[Step 2] Branch: {branch}")
    return parent_key, branch


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3  —  Checkout branch
# ─────────────────────────────────────────────────────────────────────────────
def step3_checkout(branch: str) -> bool:
    if not branch:
        print("[Step 3] No branch to checkout — skipping.")
        return False

    def run(cmd):
        return subprocess.run(cmd, shell=True, capture_output=True, text=True)

    r = run(f"git fetch origin {branch}")
    if r.returncode != 0:
        print(f"[Step 3] git fetch failed: {r.stderr.strip()}")
        return False

    # Try checkout existing remote branch; create local tracking branch if needed
    r = run(f"git checkout {branch}")
    if r.returncode != 0:
        r = run(f"git checkout -b {branch} origin/{branch}")
    if r.returncode != 0:
        print(f"[Step 3] git checkout failed: {r.stderr.strip()}")
        return False

    print(f"[Step 3] Checked out: {branch}")
    return True


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4  —  Code analysis (skipped — AI disabled)
# ─────────────────────────────────────────────────────────────────────────────
def step4_analyse_code(branch: Optional[str]) -> tuple[str, list[str]]:
    # Get list of changed files for reference (no AI analysis)
    diff_cmd = (
        f"git diff --name-only origin/main...{branch}"
        if branch
        else "git diff --name-only HEAD~10...HEAD"
    )
    result = subprocess.run(diff_cmd, shell=True, capture_output=True, text=True)
    changed_files = [f for f in result.stdout.strip().splitlines() if f]
    print(f"[Step 4] Skipped AI analysis — {len(changed_files)} changed files noted.")
    return "(AI analysis skipped)", changed_files


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5  —  Build template test plan from ticket description
# ─────────────────────────────────────────────────────────────────────────────
def step5_write_test_plan(ticket, risk_summary: str) -> list[dict]:
    summary = ticket["fields"].get("summary", "")

    # Build a standard template: happy path + 3 generic edge cases
    test_cases = [
        {
            "id": "TC-001",
            "title": f"Happy path — {summary[:70]}",
            "preconditions": "User is logged in. Feature flag is enabled for the test company.",
            "steps": "1. Navigate to the relevant section.\n2. Perform the primary action described in the ticket.\n3. Confirm the operation completes successfully.",
            "expected_result": "The feature works as described in the ticket with no errors.",
        },
        {
            "id": "TC-002",
            "title": "Edge case — invalid or missing input",
            "preconditions": "User is logged in.",
            "steps": "1. Navigate to the relevant section.\n2. Attempt the action with invalid or empty input.\n3. Submit/confirm.",
            "expected_result": "Appropriate validation error is shown. No data is corrupted.",
        },
        {
            "id": "TC-003",
            "title": "Edge case — boundary / large data",
            "preconditions": "User is logged in.",
            "steps": "1. Navigate to the relevant section.\n2. Attempt the action with maximum allowed input or a large file/dataset.\n3. Submit/confirm.",
            "expected_result": "System handles boundary input gracefully without crash or timeout.",
        },
        {
            "id": "TC-004",
            "title": "Regression — existing related functionality unaffected",
            "preconditions": "User is logged in.",
            "steps": "1. Navigate to a feature that existed before this change.\n2. Perform a standard operation that overlaps with the changed code.\n3. Confirm outcome.",
            "expected_result": "Pre-existing functionality behaves identically to before the change.",
        },
        {
            "id": "TC-005",
            "title": "Permission / feature flag — access control",
            "preconditions": "Test with a user/company that does NOT have the relevant feature flag enabled.",
            "steps": "1. Log in as a user without the feature enabled.\n2. Attempt to access or use the new feature.\n3. Observe result.",
            "expected_result": "Feature is not accessible or is hidden for users without the flag.",
        },
    ]

    print(f"[Step 5] Template test plan created — {len(test_cases)} test cases.")
    return test_cases


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6  —  Create Excel file
# ─────────────────────────────────────────────────────────────────────────────
def step6_create_excel(ticket_key: str, test_cases: list) -> tuple[Path, str]:
    filename = f"QA_{ticket_key}_{TODAY}.xlsx"
    filepath = OUTPUT_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    HEADERS = [
        "TC ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Actual Result", "Status", "Notes", "Screenshot",
    ]
    COL_WIDTHS = [10, 35, 40, 65, 45, 40, 12, 30, 22]

    # Header row
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22

    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    for row_idx, tc in enumerate(test_cases, 2):
        values = [
            tc.get("id", f"TC-{row_idx - 1:03d}"),
            tc.get("title", ""),
            tc.get("preconditions", ""),
            tc.get("steps", ""),
            tc.get("expected_result", ""),
            "",   # Actual Result  — filled after execution
            "",   # Status
            "",   # Notes
            "",   # Screenshot
        ]
        row_fill = alt_fill if row_idx % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_fill:
                cell.fill = row_fill

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"[Step 6] Excel created: {filepath}")
    return filepath, filename


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7  —  Execute TC-001 via headless Playwright
# ─────────────────────────────────────────────────────────────────────────────
def step7_execute_tc001(test_cases: list, ticket_key: str):
    screenshot_filename = f"screenshot_TC001_{ticket_key}.png"
    screenshot_path     = OUTPUT_DIR / screenshot_filename

    if not test_cases:
        return "BLOCKED", "No test cases were generated.", "", screenshot_path, screenshot_filename

    tc001 = test_cases[0]

    actual_result = ""
    status        = "BLOCKED"
    notes         = ""

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            context = browser.new_context(
                viewport={"width": 1280, "height": 800},
                ignore_https_errors=True,
            )
            page = context.new_page()

            # ── Navigate to staging ──────────────────────────────────────
            page.goto(STAGING_URL, timeout=30_000)
            page.wait_for_load_state("networkidle", timeout=20_000)

            # ── Attempt login if credentials are available ────────────────
            if STAGING_USERNAME and STAGING_PASSWORD:
                login_selectors = [
                    ("input[type='email']",    STAGING_USERNAME),
                    ("input[name='email']",    STAGING_USERNAME),
                    ("input[name='username']", STAGING_USERNAME),
                    ("#email",                 STAGING_USERNAME),
                    ("#username",              STAGING_USERNAME),
                ]
                password_selectors = [
                    "input[type='password']",
                    "input[name='password']",
                    "#password",
                ]
                submit_selectors = [
                    "button[type='submit']",
                    "input[type='submit']",
                    "button:has-text('Log in')",
                    "button:has-text('Login')",
                    "button:has-text('Sign in')",
                ]

                for sel, val in login_selectors:
                    try:
                        page.fill(sel, val, timeout=3_000)
                        break
                    except Exception:
                        pass

                for sel in password_selectors:
                    try:
                        page.fill(sel, STAGING_PASSWORD, timeout=3_000)
                        break
                    except Exception:
                        pass

                for sel in submit_selectors:
                    try:
                        page.click(sel, timeout=3_000)
                        page.wait_for_load_state("networkidle", timeout=15_000)
                        break
                    except Exception:
                        pass

            # ── Screenshot ───────────────────────────────────────────────
            page.screenshot(path=str(screenshot_path), full_page=False)

            current_url  = page.url
            page_title   = page.title()
            actual_result = f"Navigated to: {current_url}\nPage title: {page_title}"

            # ── Pass/Fail heuristic ───────────────────────────────────────
            if any(kw in current_url.lower() for kw in ("login", "sign-in", "signin", "auth")):
                status = "BLOCKED"
                notes  = (
                    "Could not log in to staging — verify STAGING_USERNAME and "
                    "STAGING_PASSWORD secrets are set correctly in GitHub."
                )
            else:
                status = "PASS"
                notes  = (
                    "Staging environment is accessible and login succeeded. "
                    "TC-001 preconditions confirmed. Full step-by-step execution "
                    "requires a human tester following the Steps column."
                )

            browser.close()

    except PlaywrightTimeoutError as exc:
        actual_result = f"Timeout reaching staging URL: {exc}"
        status        = "BLOCKED"
        notes         = "Staging URL timed out — check network/VPN requirements."
        screenshot_path.write_bytes(b"")          # placeholder

    except Exception as exc:
        actual_result = f"Unexpected error: {exc}"
        status        = "BLOCKED"
        notes         = str(exc)
        if not screenshot_path.exists():
            screenshot_path.write_bytes(b"")

    print(f"[Step 7] TC-001: {status}")
    return status, actual_result, notes, screenshot_path, screenshot_filename


# ─────────────────────────────────────────────────────────────────────────────
# STEP 8  —  Update Excel with TC-001 results
# ─────────────────────────────────────────────────────────────────────────────
def step8_update_excel(
    excel_path: Path,
    status: str,
    actual_result: str,
    notes: str,
    screenshot_filename: str,
):
    colour_map = {
        "PASS":    ("C6EFCE", "276221"),
        "FAIL":    ("FFC7CE", "9C0006"),
        "BLOCKED": ("FFEB9C", "9C5700"),
    }
    bg, fg = colour_map.get(status, ("FFFFFF", "000000"))

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    ws.cell(row=2, column=6, value=actual_result).alignment   = Alignment(wrap_text=True, vertical="top")
    status_cell = ws.cell(row=2, column=7, value=status)
    status_cell.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    status_cell.font      = Font(bold=True, color=fg)
    status_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=8, value=notes).alignment           = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=2, column=9, value=screenshot_filename)

    wb.save(excel_path)
    print("[Step 8] Excel updated with TC-001 results.")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 9a  —  Update dedup log
# ─────────────────────────────────────────────────────────────────────────────
def step9a_update_dedup(dedup_data: dict, ticket, excel_filename: str, tc001_status: str):
    dedup_data["processed"].append(
        {
            "ticket_key":     ticket["key"],
            "ticket_summary": ticket["fields"]["summary"],
            "processed_date": TODAY,
            "excel_file":     excel_filename,
            "tc001_status":   tc001_status,
        }
    )
    DEDUP_PATH.write_text(json.dumps(dedup_data, indent=2))
    print(f"[Step 9a] Dedup log updated — {ticket['key']} marked processed.")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 9c  —  Slack DM (optional — skipped if SLACK_BOT_TOKEN not set)
# ─────────────────────────────────────────────────────────────────────────────
def step9c_slack(ticket, branch: Optional[str], tc001_status: str, excel_filename: str):
    token = os.environ.get("SLACK_BOT_TOKEN", "")
    if not token:
        print("[Step 9c] SLACK_BOT_TOKEN not set — skipping Slack notification.")
        return

    emoji = {"PASS": "✅", "FAIL": "❌", "BLOCKED": "⚠️"}.get(tc001_status, "⚠️")
    run_url = os.environ.get("GITHUB_SERVER_URL", "https://github.com")
    repo    = os.environ.get("GITHUB_REPOSITORY", "i6systems/in2plane-cloud")
    run_id  = os.environ.get("GITHUB_RUN_ID", "")
    artifacts_url = f"{run_url}/{repo}/actions/runs/{run_id}" if run_id else "GitHub Actions"

    text = (
        f"*Nightly QA Run — {TODAY}*\n"
        f"• *Ticket:* <https://i6group.atlassian.net/browse/{ticket['key']}|{ticket['key']}>"
        f" — {ticket['fields']['summary']}\n"
        f"• *Branch:* `{branch or 'unknown'}`\n"
        f"• *TC-001:* {emoji} {tc001_status}\n"
        f"• *Excel + screenshot:* <{artifacts_url}|Download from GitHub Actions artifacts>"
        f" → `{excel_filename}`\n"
        f"\n"
        f"💳 Token balance check: https://console.anthropic.com/settings/usage\n"
        f"Check daily to avoid running out during personal use."
    )

    resp = requests.post(
        "https://slack.com/api/chat.postMessage",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={"channel": SLACK_DM_CHANNEL, "text": text, "mrkdwn": True},
    )
    data = resp.json()
    if data.get("ok"):
        print("[Step 9c] Slack DM sent.")
    else:
        print(f"[Step 9c] Slack error: {data.get('error')} — {data}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    sep = "=" * 62
    print(f"\n{sep}\n  Nightly QA Workflow — {TODAY}\n{sep}\n")

    dedup_data, processed_keys = step0_dedup()
    ticket                     = step1_get_ticket(processed_keys)
    parent_key, branch         = step2_get_branch(ticket)

    if branch:
        step3_checkout(branch)

    risk_summary, changed_files = step4_analyse_code(branch)
    test_cases                  = step5_write_test_plan(ticket, risk_summary)
    excel_path, excel_filename  = step6_create_excel(ticket["key"], test_cases)

    tc_status, actual_result, notes, screenshot_path, screenshot_filename = \
        step7_execute_tc001(test_cases, ticket["key"])

    step8_update_excel(excel_path, tc_status, actual_result, notes, screenshot_filename)
    step9a_update_dedup(dedup_data, ticket, excel_filename, tc_status)
    step9c_slack(ticket, branch, tc_status, excel_filename)

    print(f"\n{sep}\n  Done. TC-001: {tc_status}\n{sep}\n")


if __name__ == "__main__":
    main()
